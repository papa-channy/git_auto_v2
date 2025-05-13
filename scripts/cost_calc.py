import json
import math
import pandas as pd
import tiktoken
import requests
from bs4 import BeautifulSoup
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.path import get_cost_log_path, get_result_paths

# ─────────────────────────────────────────────
# 환율 처리 (하루 1회만 크롤링)
# ─────────────────────────────────────────────
def get_usd_exchange_rate(log) -> float:
    ex_file = Path("cost/ex_rate.txt")
    today = datetime.now().strftime("%Y-%m-%d")

    if ex_file.exists():
        last_modified = datetime.fromtimestamp(ex_file.stat().st_mtime).strftime("%Y-%m-%d")
        if last_modified == today:
            log("💱 환율 캐시 사용")
            return float(ex_file.read_text().strip())

    try:
        res = requests.get("https://finance.naver.com/marketindex/")
        soup = BeautifulSoup(res.text, "html.parser")
        price = soup.select_one("div.head_info > span.value").text
        rate = float(price.replace(",", ""))
        ex_file.write_text(str(rate))
        log(f"📈 실시간 환율 업데이트: {rate} 원/USD")
        return rate
    except:
        log("⚠️ 환율 크롤링 실패, 이전 환율 또는 기본값 사용")
        return float(ex_file.read_text().strip()) if ex_file.exists() else 1400.0

# ─────────────────────────────────────────────
# 토큰 계산기
# ─────────────────────────────────────────────
def count_tokens(text: str, model: str) -> int:
    try:
        enc = tiktoken.encoding_for_model(model)
    except:
        enc = tiktoken.get_encoding("cl100k_base")
    return len(enc.encode(text))

# ─────────────────────────────────────────────
# 엑셀 저장기 (append)
# ─────────────────────────────────────────────
def append_to_excel(df: pd.DataFrame, timestamp: str, filepath: str, log):
    dt = datetime.strptime(timestamp, "%Y%m%d_%H%M")
    df.insert(0, "datetime", dt)

    file = Path(filepath)
    if file.exists():
        wb = load_workbook(file)
        ws = wb.active
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
    else:
        wb = Workbook()
        ws = wb.active
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    wb.save(file)
    log(f"✅ 엑셀 누적 저장 완료 → {filepath}")

# ─────────────────────────────────────────────
# 메인 실행 함수
# ─────────────────────────────────────────────
def calculate_llm_costs(prompt_dict: dict, timestamp: str, log):
    rate_config = json.loads(Path("config/cost.json").read_text(encoding="utf-8"))
    krw_rate = get_usd_exchange_rate(log)
    results = []

    result_paths = get_result_paths(timestamp)
    txt_sources = {
        "a": result_paths["context_by_file"].glob("*.txt"),
        "b": [result_paths["context_summary"]],
        "c": result_paths["diff_chunks"].glob("*.txt"),
        "d": [result_paths["diff_final"]],
        "e": [result_paths["final_record"]],
    }

    outputs = {}
    for key, files in txt_sources.items():
        for i, f in enumerate(sorted(files)):
            var = f"{key}{i+1}" if key in {"a", "c"} else key
            if f.exists():
                outputs[f"{var}_ans"] = f.read_text(encoding="utf-8")

    for var, meta in prompt_dict.items():
        text = meta["text"]
        model = meta["model"]
        purpose = meta["purpose"]

        input_tokens = count_tokens(text, model)
        output_text = outputs.get(f"{var}_ans", "")
        output_tokens = count_tokens(output_text, model) if output_text else 0

        in_rate = rate_config[model]["input"]
        out_rate = rate_config[model]["output"]

        usd = (input_tokens * in_rate + output_tokens * out_rate)
        krw = usd * krw_rate

        results.append({
            "var": var,
            "purpose": purpose,
            "model": model,
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "usd": round(usd, 6),
            "krw": round(krw),
        })

    df = pd.DataFrame(results)
    df = df[["datetime", "var", "purpose", "model", "input_tokens", "output_tokens", "usd", "krw"]]

    append_to_excel(df, timestamp, "cost/git_cost.xlsx", log)
    return df
